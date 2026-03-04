"""
Exportación F10-03 (Validación producto): workbook generado desde dict de solicitud, con formato legible.
"""

from __future__ import annotations

import io
import re
from copy import copy
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SECTION_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _producto_nombre(solicitud: dict[str, Any]) -> str:
    f01 = solicitud.get("f10_01") or {}
    return (f01.get("NOMBRE") or f01.get("NOM_COMERCIAL") or "").strip() or "—"


def _responsable(solicitud: dict[str, Any]) -> str:
    f02 = solicitud.get("f10_02") or {}
    return (f02.get("responsable") or "").strip()


# Caracteres no válidos en nombre de hoja Excel: \ / ? * [ ]
_SHEET_NAME_BAD = re.compile(r'[\\/:*?\[\]]')


def _sheet_name_for_solicitud(solicitud: dict[str, Any], max_len: int = 31) -> str:
    """Nombre de hoja: Nº Solicitud normalizado con '-' + Nombre proyecto. Máx 31 caracteres."""
    num = (solicitud.get("numero_solicitud") or "").strip().replace(" ", "-").replace("/", "-")
    if num and num != "-":
        num = num.strip("-")
    proy = _producto_nombre(solicitud)
    if num and proy and proy != "—":
        raw = f"{num} {proy}"
    else:
        raw = num or proy or "solicitud"
    raw = _SHEET_NAME_BAD.sub("_", raw)[:max_len].strip()
    return raw or "Hoja"


def _copy_sheet_content(source: Worksheet, target: Worksheet) -> None:
    """Copia celdas (valor y estilo) de source a target (pueden ser de distintos workbooks)."""
    for row in source.iter_rows():
        for cell in row:
            t = target.cell(row=cell.row, column=cell.column)
            t.value = cell.value
            if cell.font:
                t.font = copy(cell.font)
            if cell.fill:
                t.fill = copy(cell.fill)
            if cell.alignment:
                t.alignment = copy(cell.alignment)
            if cell.border:
                t.border = copy(cell.border)
            if cell.number_format:
                t.number_format = cell.number_format
    for col_letter, col_dim in source.column_dimensions.items():
        if col_dim.width is not None:
            target.column_dimensions[col_letter].width = col_dim.width


def _parse_caracteristicas_quimicas(text: str) -> List[tuple[str, str]]:
    """Convierte el texto 'Param\\tValor' por línea en lista de (parámetro, valor)."""
    if not (text and str(text).strip()):
        return []
    rows = []
    for line in str(text).strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if "\t" in line:
            parts = line.split("\t", 1)
        else:
            parts = line.split(None, 1)  # primer espacio
        param = (parts[0] or "").strip()
        valor = (parts[1] or "").strip() if len(parts) > 1 else ""
        rows.append((param, valor))
    return rows


def build_f10_03_workbook(solicitud: dict[str, Any]) -> Workbook:
    """Construye un workbook F10-03 rellenado desde el dict (f10_03, f10_01, f10_02)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "F10-03 Validación producto"

    f03 = solicitud.get("f10_03") or {}
    esp = f03.get("especificacion_final") or {}
    val = f03.get("validacion") or {}
    filas: List[Dict[str, Any]] = val.get("filas") or []

    numero = solicitud.get("numero_solicitud") or ""
    producto = _producto_nombre(solicitud)
    responsable = _responsable(solicitud)

    row = 1

    # Título
    ws.cell(row, 1, "F10-03 — Validación de producto")
    ws.cell(row, 1).font = TITLE_FONT
    row += 2

    # Identificación
    ws.cell(row, 1, "Nº Solicitud:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, str(numero))
    row += 1
    ws.cell(row, 1, "Responsable:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, str(responsable))
    row += 1
    ws.cell(row, 1, "Producto:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, producto)
    row += 2

    # 1. ESPECIFICACIÓN FINAL
    ws.cell(row, 1, "1. ESPECIFICACIÓN FINAL")
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 1
    ws.cell(row, 1, "Descripción")
    ws.cell(row, 1).font = LABEL_FONT
    desc_cell = ws.cell(row, 2, esp.get("descripcion", ""))
    desc_cell.alignment = WRAP_ALIGNMENT
    row += 2
    ws.cell(row, 1, "Aspecto")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, esp.get("aspecto", ""))
    row += 1
    ws.cell(row, 1, "Color")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, esp.get("color", ""))
    row += 2
    # Características químicas: tabla de dos columnas (parámetro | valor), una fila por parámetro
    car_quim_text = esp.get("caracteristicas_quimicas", "") or ""
    car_quim_rows = _parse_caracteristicas_quimicas(car_quim_text)
    ws.cell(row, 1, "Características químicas")
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 1
    # Cabecera tabla
    ws.cell(row, 1, "Parámetro")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.cell(row, 1).border = THIN_BORDER
    ws.cell(row, 2, "Valor")
    ws.cell(row, 2).font = LABEL_FONT
    ws.cell(row, 2).fill = SECTION_FILL
    ws.cell(row, 2).border = THIN_BORDER
    row += 1
    for param, valor in car_quim_rows:
        ws.cell(row, 1, param).border = THIN_BORDER
        ws.cell(row, 2, valor).border = THIN_BORDER
        row += 1
    row += 1

    # 2. VALIDACIÓN
    ws.cell(row, 1, "2. VALIDACIÓN")
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 1
    ws.cell(row, 1, "Fecha validación")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, val.get("fecha_validacion", ""))
    row += 2
    # Tabla: Área, Aspecto a validar, OK/NOK, Comentarios
    headers = ["Área", "Aspecto a validar", "OK/NOK", "Comentarios"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row, col, h)
        cell.font = LABEL_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER
    row += 1
    for f in filas:
        ws.cell(row, 1, str(f.get("area", ""))).border = THIN_BORDER
        ws.cell(row, 2, str(f.get("aspecto_a_validar", ""))).border = THIN_BORDER
        ws.cell(row, 3, str(f.get("validar_ok_nok", ""))).border = THIN_BORDER
        com_cell = ws.cell(row, 4, str(f.get("comentarios", "")))
        com_cell.border = THIN_BORDER
        com_cell.alignment = WRAP_ALIGNMENT
        row += 1

    # Anchos: ~20, ~35, ~10, ~25
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 25
    return wb


def build_f10_03_workbook_all(solicitudes: List[dict[str, Any]]) -> Workbook:
    """Un workbook con una hoja por solicitud; nombre de hoja = Nº Solicitud normalizado + Nombre proyecto."""
    if not solicitudes:
        wb = Workbook()
        wb.active.title = "Vacío"
        return wb
    wb_main = Workbook()
    seen_names: set[str] = set()
    for idx, sol in enumerate(solicitudes):
        if not isinstance(sol, dict):
            continue
        wb_one = build_f10_03_workbook(sol)
        ws_one = wb_one.active
        base_name = _sheet_name_for_solicitud(sol)
        sheet_name = base_name
        suffix = 0
        while sheet_name in seen_names:
            suffix += 1
            sheet_name = f"{base_name[:28]}_{suffix}" if len(base_name) >= 28 else f"{base_name}_{suffix}"
        seen_names.add(sheet_name)
        if idx == 0:
            wb_main.active.title = sheet_name
            _copy_sheet_content(ws_one, wb_main.active)
        else:
            new_ws = wb_main.create_sheet(sheet_name)
            _copy_sheet_content(ws_one, new_ws)
    return wb_main


def build_f10_03_bytes_all(solicitudes: List[dict[str, Any]]) -> bytes:
    """Devuelve el workbook F10-03 (todas las solicitudes, una hoja por solicitud) como bytes."""
    wb = build_f10_03_workbook_all(solicitudes)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_f10_03_bytes(solicitud: dict[str, Any]) -> bytes:
    """Devuelve el workbook F10-03 como bytes para descarga."""
    wb = build_f10_03_workbook(solicitud)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
