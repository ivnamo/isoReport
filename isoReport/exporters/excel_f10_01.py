"""
Exportación F10-01 (Viabilidad y planificación): una hoja con todas las solicitudes, una fila por solicitud.
Estructura y columnas alineadas al CSV de referencia; formato legible para auditoría.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Orden de columnas igual que F10-01 Viabilidad y planificación de diseños__2025.csv
F10_01_COLUMNS = [
    "ID",
    "Nº Solicitud",
    "SOLICITANTE",
    "NOMBRE",
    "NOM_COMERCIAL",
    "NECESIDAD",
    "VOLUMEN COMPETENCIA",
    "PRECIO COMPETENCIA",
    "VOLUMEN PROPUESTO",
    "ENVASE",
    "PAIS DESTINO",
    "ACEPTADO",
    "FINALIZADO",
    "MOTIVO DENEGADO",
    "FECHA DE APROBACIÓN SOL.",
    "TIEMPO ESTIMADO (días laborables)",
    "FECHA FINALIZACION ESTIMADA",
    "FECHA DE FINALIZACION REAL",
    "HORAS EMPLEADAS I+D",
    "HORAS EMPLEADAS EN CALIDAD",
    "PROBLEMAS",
    "COMENTARIOS",
]

# Anchos por columna (índice 0-based)
F10_01_WIDTHS = {
    0: 6,   # ID
    1: 12,  # Nº Solicitud
    2: 10,  # SOLICITANTE
    3: 28,  # NOMBRE
    4: 28,  # NOM_COMERCIAL
    5: 55,  # NECESIDAD
    6: 18,  # VOLUMEN COMPETENCIA
    7: 18,  # PRECIO COMPETENCIA
    8: 18,  # VOLUMEN PROPUESTO
    9: 12,  # ENVASE
    10: 15, # PAIS DESTINO
    11: 10, # ACEPTADO
    12: 10, # FINALIZADO
    13: 25, # MOTIVO DENEGADO
    14: 18, # FECHA DE APROBACIÓN SOL.
    15: 18, # TIEMPO ESTIMADO
    16: 18, # FECHA FINALIZACION ESTIMADA
    17: 18, # FECHA DE FINALIZACION REAL
    18: 18, # HORAS EMPLEADAS I+D
    19: 22, # HORAS EMPLEADAS EN CALIDAD
    20: 60, # PROBLEMAS
    21: 25, # COMENTARIOS
}

WRAP_COLUMN_INDICES = {5, 13, 20, 21}  # NECESIDAD, MOTIVO DENEGADO, PROBLEMAS, COMENTARIOS

HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _row_from_solicitud(solicitud: dict[str, Any]) -> list[Any]:
    f01 = solicitud.get("f10_01") or {}
    row = []
    for col in F10_01_COLUMNS:
        if col == "ID":
            val = solicitud.get("id")
        else:
            val = f01.get(col, "")
        row.append("" if val is None else str(val).strip())
    return row


def build_f10_01_workbook(solicitudes: list[dict[str, Any]]) -> Workbook:
    """Construye un workbook F10-01: una hoja, una fila por solicitud, formato legible."""
    wb = Workbook()
    ws = wb.active
    ws.title = "F10-01 Viabilidad"

    # Cabecera
    for col_idx, header in enumerate(F10_01_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Filas de datos
    for row_idx, solicitud in enumerate(solicitudes, start=2):
        if not isinstance(solicitud, dict):
            continue
        values = _row_from_solicitud(solicitud)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx - 1 in WRAP_COLUMN_INDICES:
                cell.alignment = WRAP_ALIGNMENT

    # Anchos de columna
    for col_idx in range(1, len(F10_01_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = F10_01_WIDTHS.get(col_idx - 1, 15)

    ws.freeze_panes = "A2"

    return wb


def build_f10_01_bytes(solicitudes: list[dict[str, Any]]) -> bytes:
    """Devuelve el workbook F10-01 como bytes para descarga."""
    wb = build_f10_01_workbook(solicitudes)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
