"""
Exportación F10-02 (Diseño producto): workbook generado desde dict de solicitud, con formato legible.
"""

from __future__ import annotations

import io
import re
from copy import copy
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SECTION_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _producto_nombre(solicitud: dict[str, Any]) -> str:
    f01 = solicitud.get("f10_01") or {}
    return (f01.get("NOMBRE") or f01.get("NOM_COMERCIAL") or "").strip() or "—"


# Caracteres no válidos en nombre de hoja Excel: \ / ? * [ ]
_SHEET_NAME_BAD = re.compile(r'[\\/:*?\[\]]')


def _sheet_name_for_solicitud(solicitud: dict[str, Any], max_len: int = 31) -> str:
    """Nombre de hoja: Nº Solicitud normalizado con '-' + Nombre proyecto. Máx 31 caracteres."""
    num = (solicitud.get("numero_solicitud") or "").strip().replace(" ", "-").replace("/", "-")
    if num and num != "-":
        num = num.strip("-")
    proy = _producto_nombre(solicitud)
    if num and proy and proy != "—":
        raw = f"{num} {proy}"
    else:
        raw = num or proy or "solicitud"
    raw = _SHEET_NAME_BAD.sub("_", raw)[:max_len].strip()
    return raw or "Hoja"


def _copy_sheet_content(source: Worksheet, target: Worksheet) -> None:
    """Copia celdas (valor y estilo) de source a target (pueden ser de distintos workbooks)."""
    for row in source.iter_rows():
        for cell in row:
            t = target.cell(row=cell.row, column=cell.column)
            t.value = cell.value
            if cell.font:
                t.font = copy(cell.font)
            if cell.fill:
                t.fill = copy(cell.fill)
            if cell.alignment:
                t.alignment = copy(cell.alignment)
            if cell.border:
                t.border = copy(cell.border)
            if cell.number_format:
                t.number_format = cell.number_format
    for col_letter, col_dim in source.column_dimensions.items():
        if col_dim.width is not None:
            target.column_dimensions[col_letter].width = col_dim.width


def build_f10_02_workbook(solicitud: dict[str, Any]) -> Workbook:
    """Construye un workbook F10-02 rellenado desde el dict de la solicitud (f10_02, f10_01)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "F10-02 Diseño producto"

    f02 = solicitud.get("f10_02") or {}
    v = f02.get("verificacion_diseno") or {}
    numero = solicitud.get("numero_solicitud") or ""
    producto = _producto_nombre(solicitud)
    responsable = f02.get("responsable", "")

    row = 1

    # Título
    ws.cell(row, 1, "F10-02 — Diseño de producto")
    ws.cell(row, 1).font = TITLE_FONT
    row += 2

    # Identificación
    ws.cell(row, 1, "Nº Solicitud:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, str(numero))
    row += 1
    ws.cell(row, 1, "Responsable:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, str(responsable))
    row += 1
    ws.cell(row, 1, "Producto:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, producto)
    row += 2

    # 1. DATOS DE PARTIDA
    ws.cell(row, 1, "1. DATOS DE PARTIDA DEL DISEÑO")
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 1
    desc_cell = ws.cell(row, 1, f02.get("descripcion_partida_diseno", ""))
    desc_cell.alignment = WRAP_ALIGNMENT
    row += 2

    # 2. ENSAYOS
    ws.cell(row, 1, "2. ENSAYOS / FORMULACIÓN")
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 2
    ensayos = f02.get("ensayos") or []
    for idx, ens in enumerate(ensayos, start=1):
        row = _write_ensayo_block(ws, ens, idx, row)
        row += 2

    # 3. VERIFICACIÓN
    ws.cell(row, 1, "3. VERIFICACIÓN (DISEÑO)")
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 1
    ws.cell(row, 1, "Producto final:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, v.get("producto_final", ""))
    row += 1
    ws.cell(row, 1, "Fórmula OK:")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 2, v.get("formula_ok", ""))
    row += 1
    ws.cell(row, 1, "Riquezas:")
    ws.cell(row, 1).font = LABEL_FONT
    riq_cell = ws.cell(row, 2, v.get("riquezas", ""))
    riq_cell.alignment = WRAP_ALIGNMENT
    row += 1

    # Anchos
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 18
    return wb


def _write_ensayo_block(ws: Worksheet, ens: Dict[str, Any], idx: int, start_row: int) -> int:
    row = start_row
    ws.cell(row, 1, f"Ensayo {idx}")
    ws.cell(row, 1).font = LABEL_FONT
    row += 1
    # Cabecera tabla identificación
    for col, (label, key) in enumerate(
        [("ID", "id"), ("Nombre", "ensayo"), ("Fecha", "fecha"), ("Resultado", "resultado")], start=1
    ):
        cell = ws.cell(row, col, label)
        cell.font = LABEL_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER
    row += 1
    for col, key in enumerate(["id", "ensayo", "fecha", "resultado"], start=1):
        ws.cell(row, col, str(ens.get(key, ""))).border = THIN_BORDER
    row += 1
    # Tabla fórmula: Materia prima | % peso
    ws.cell(row, 1, "Materia prima")
    ws.cell(row, 1).font = LABEL_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.cell(row, 1).border = THIN_BORDER
    ws.cell(row, 2, "% peso")
    ws.cell(row, 2).font = LABEL_FONT
    ws.cell(row, 2).fill = SECTION_FILL
    ws.cell(row, 2).border = THIN_BORDER
    row += 1
    formula = ens.get("formula") or []
    for item in formula:
        ws.cell(row, 1, str(item.get("materia_prima", ""))).border = THIN_BORDER
        ws.cell(row, 2, str(item.get("porcentaje_peso", ""))).border = THIN_BORDER
        row += 1
    row += 1
    ws.cell(row, 1, "Motivo / comentario")
    ws.cell(row, 1).font = LABEL_FONT
    mot_cell = ws.cell(row, 2, str(ens.get("motivo_comentario", "")))
    mot_cell.alignment = WRAP_ALIGNMENT
    row += 1
    return row


def build_f10_02_workbook_all(solicitudes: List[dict[str, Any]]) -> Workbook:
    """Un workbook con una hoja por solicitud; nombre de hoja = Nº Solicitud normalizado + Nombre proyecto."""
    if not solicitudes:
        wb = Workbook()
        wb.active.title = "Vacío"
        return wb
    wb_main = Workbook()
    seen_names: set[str] = set()
    for idx, sol in enumerate(solicitudes):
        if not isinstance(sol, dict):
            continue
        wb_one = build_f10_02_workbook(sol)
        ws_one = wb_one.active
        base_name = _sheet_name_for_solicitud(sol)
        sheet_name = base_name
        suffix = 0
        while sheet_name in seen_names:
            suffix += 1
            sheet_name = f"{base_name[:28]}_{suffix}" if len(base_name) >= 28 else f"{base_name}_{suffix}"
        seen_names.add(sheet_name)
        if idx == 0:
            wb_main.active.title = sheet_name
            _copy_sheet_content(ws_one, wb_main.active)
        else:
            new_ws = wb_main.create_sheet(sheet_name)
            _copy_sheet_content(ws_one, new_ws)
    return wb_main


def build_f10_02_bytes_all(solicitudes: List[dict[str, Any]]) -> bytes:
    """Devuelve el workbook F10-02 (todas las solicitudes, una hoja por solicitud) como bytes."""
    wb = build_f10_02_workbook_all(solicitudes)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_f10_02_bytes(solicitud: dict[str, Any]) -> bytes:
    """Devuelve el workbook F10-02 como bytes para descarga."""
    wb = build_f10_02_workbook(solicitud)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
