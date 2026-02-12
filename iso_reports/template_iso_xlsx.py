from __future__ import annotations

import io
from typing import List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import Ensayo, InformeData


def create_iso_workbook(informe: InformeData) -> Workbook:
    """
    Crea un Workbook de Excel con un layout similar al informe
    ISO actual (incluyendo el anexo), rellenado con los datos
    de `informe`.

    No depende de una plantilla externa: define el formato mínimo
    mediante código para facilitar el despliegue en Streamlit Cloud.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "INFORME_ISO"

    current_row = 1

    # Cabecera
    ws.cell(current_row, 1, "Responsable de proyecto:")
    ws.cell(current_row, 2, informe.responsable)
    current_row += 1

    ws.cell(current_row, 1, "Nº Solicitud:")
    ws.cell(current_row, 2, informe.numero_solicitud)
    ws.cell(current_row, 3, "Tipo:")
    ws.cell(current_row, 4, informe.tipo_solicitud)
    current_row += 1

    ws.cell(current_row, 1, "Producto base / línea:")
    ws.cell(current_row, 2, informe.producto_base)
    current_row += 2  # línea en blanco

    # 1. Datos de partida
    ws.cell(current_row, 1, "1. DATOS DE PARTIDA DEL DISEÑO")
    current_row += 1
    ws.cell(current_row, 1, informe.descripcion_diseno)
    current_row += 2

    # 2. Ensayos / formulaciones
    ws.cell(current_row, 1, "2. ENSAYOS / FORMULACIONES")
    current_row += 2

    for idx, ensayo in enumerate(informe.ensayos, start=1):
        current_row = _write_ensayo_block(ws, ensayo, idx, start_row=current_row)
        current_row += 2  # separación entre ensayos

    # 3. Verificación
    ws.cell(current_row, 1, "3. VERIFICACIÓN")
    current_row += 1
    ws.cell(current_row, 1, "Producto final:")
    ws.cell(current_row, 2, informe.producto_final)
    current_row += 1
    ws.cell(current_row, 1, "Fórmula OK:")
    ws.cell(current_row, 2, informe.formula_ok)
    current_row += 1
    ws.cell(current_row, 1, "Riquezas:")
    ws.cell(current_row, 2, informe.riquezas)
    current_row += 2

    # ANEXO F90-02
    espec = informe.especificacion_final
    valid = informe.validacion_producto

    ws.cell(current_row, 1, "ANEXO F90-02: VALIDACIÓN DE PRODUCTO")
    current_row += 1

    ws.cell(current_row, 1, "1. ESPECIFICACIÓN FINAL")
    current_row += 1

    ws.cell(current_row, 1, "DESCRIPCIÓN")
    ws.cell(current_row, 2, espec.descripcion)
    current_row += 2

    ws.cell(current_row, 1, "CARACTERÍSTICAS FÍSICAS")
    current_row += 1
    ws.cell(current_row, 1, "Aspecto")
    ws.cell(current_row, 2, espec.aspecto)
    current_row += 1
    ws.cell(current_row, 1, "Densidad")
    ws.cell(current_row, 2, espec.densidad)
    current_row += 1
    ws.cell(current_row, 1, "Color")
    ws.cell(current_row, 2, espec.color)
    current_row += 1
    ws.cell(current_row, 1, "pH")
    ws.cell(current_row, 2, espec.ph)
    current_row += 2

    ws.cell(current_row, 1, "CARACTERÍSTICAS QUÍMICAS")
    current_row += 1
    ws.cell(current_row, 1, espec.caracteristicas_quimicas)
    current_row += 2

    ws.cell(current_row, 1, "2. VALIDACIÓN (El producto satisface los requisitos)")
    current_row += 1
    ws.cell(current_row, 1, "Fecha de validación")
    ws.cell(current_row, 2, valid.fecha_validacion)
    current_row += 1
    ws.cell(current_row, 1, "Comentario validación")
    ws.cell(current_row, 2, valid.comentario_validacion)

    # Ajuste básico de anchos de columna
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    return wb


def _write_ensayo_block(
    ws: Worksheet,
    ensayo: Ensayo,
    idx: int,
    start_row: int,
) -> int:
    """
    Escribe el bloque correspondiente a un ensayo, empezando en start_row.
    Devuelve la siguiente fila libre después del bloque.
    """
    row = start_row

    ws.cell(row, 1, f"Ensayo {idx}")
    ws.cell(row, 2, ensayo.id_ensayo)
    ws.cell(row, 3, ensayo.nombre_formulacion)
    row += 1

    # Fila fecha / resultado / datos Jira
    ws.cell(row, 1, "Fecha ensayo:")
    ws.cell(row, 2, ensayo.fecha_ensayo)
    ws.cell(row, 3, "Resultado:")
    ws.cell(row, 4, ensayo.resultado)

    if ensayo.jira_estado:
        ws.cell(row, 5, "Estado Jira:")
        ws.cell(row, 6, ensayo.jira_estado)
    if ensayo.jira_persona_asignada:
        ws.cell(row, 7, "Asignado:")
        ws.cell(row, 8, ensayo.jira_persona_asignada)
    row += 2

    # Tabla de materias primas
    ws.cell(row, 1, "Materia prima")
    ws.cell(row, 2, "% peso")
    row += 1

    for mp in ensayo.materias:
        ws.cell(row, 1, mp.nombre)
        ws.cell(row, 2, mp.porcentaje_peso)
        row += 1

    row += 1

    # Motivo / comentario + notas Jira
    texto = ensayo.motivo or ""
    if ensayo.jira_resumen:
        texto = f"{texto}\n\n[Resumen Jira] {ensayo.jira_resumen}".strip()
    if ensayo.jira_comentarios_resumen:
        texto = f"{texto}\n\n[Comentarios Jira]\n{ensayo.jira_comentarios_resumen}".strip()

    ws.cell(row, 1, "Motivo / comentario")
    ws.cell(row, 2, texto)
    row += 1

    return row


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Convierte un Workbook en bytes, listo para `st.download_button`."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


